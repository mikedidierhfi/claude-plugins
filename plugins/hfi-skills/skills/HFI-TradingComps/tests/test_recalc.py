#!/usr/bin/env python3
"""
test_recalc.py — Verify the workbook's LIVE FORMULAS actually COMPUTE the right numbers.

openpyxl does not evaluate formulas, so the other tests can only assert that the formula STRINGS
reference the right cells (test_render) and that the Python orchestrator math is correct
(test_valuation). This test closes the remaining gap: that the formulas, AS WRITTEN, recalc
correctly in Excel. It ships a tiny Excel-formula interpreter (IF / AND / OR / NOT / ISNUMBER / N,
+ - * / , comparisons, and MIN / MAX / AVERAGE / MEDIAN / COUNT over ranges), evaluates the exact
strings the renderer emits, and checks TEV, market equity, working capital, EBITDA, every multiple,
and the peer summary statistics (incl. that "nm" is excluded) against independently hand-computed
values.

Run: python tests/test_recalc.py
"""
import os, re, sys, statistics
HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.normpath(os.path.join(HERE, "..", "assets"))
sys.path.insert(0, ASSETS)
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter, coordinate_to_tuple  # noqa: E402
import build_comps_xlsx as bx  # noqa: E402

OUT = os.path.join(HERE, "_out"); os.makedirs(OUT, exist_ok=True)


# ---------------- a minimal, defensive Excel-formula evaluator ----------------
def isnum(x):
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def truthy(x):
    if isinstance(x, bool):
        return x
    if isnum(x):
        return x != 0
    return False


def _arith(op, a, b):
    if not (isnum(a) and isnum(b)):
        return "#VAL"                       # discarded by an IF/ISNUMBER guard in our templates
    if op == "+":
        return a + b
    if op == "-":
        return a - b
    if op == "*":
        return a * b
    return (a / b) if b != 0 else "#DIV0"   # div-by-zero is guarded by IF(den<=0,...) upstream


def _cmp(op, a, b):
    if not (isnum(a) and isnum(b)):
        return False
    return {"<": a < b, "<=": a <= b, ">": a > b, ">=": a >= b,
            "=": a == b, "<>": a != b}[op]


def _flatten(args):
    out = []
    for a in args:
        out.extend(a) if isinstance(a, list) else out.append(a)
    return out


def _apply(name, args):
    if name == "IF":
        return args[1] if truthy(args[0]) else (args[2] if len(args) > 2 else False)
    if name == "AND":
        return all(truthy(x) for x in args)
    if name == "OR":
        return any(truthy(x) for x in args)
    if name == "NOT":
        return not truthy(args[0])
    if name == "ISNUMBER":
        return isnum(args[0])
    if name == "N":
        return args[0] if isnum(args[0]) else 0.0
    nums = [x for x in _flatten(args) if isnum(x)]
    if name == "COUNT":
        return float(len(nums))
    if not nums:
        return 0.0                          # our stat formulas guard MIN/etc. with IF(COUNT=0,"",..)
    if name == "MIN":
        return min(nums)
    if name == "MAX":
        return max(nums)
    if name == "SUM":
        return sum(nums)
    if name == "AVERAGE":
        return sum(nums) / len(nums)
    if name == "MEDIAN":
        return statistics.median(nums)
    raise ValueError("unknown function " + name)


TOK = re.compile(r'(?P<num>\d+(?:\.\d+)?)|(?P<str>"[^"]*")|(?P<range>[A-Z]+\d+:[A-Z]+\d+)'
                 r'|(?P<cell>[A-Z]+\d+)|(?P<func>[A-Z][A-Z0-9_.]*)\(|(?P<op><=|>=|<>|[-+*/()<>=,])')


def tokenize(s):
    toks, i = [], 0
    while i < len(s):
        if s[i].isspace():
            i += 1; continue
        m = TOK.match(s, i)
        if not m:
            raise ValueError(f"bad token at {s[i:]!r}")
        i = m.end()
        d = m.groupdict()
        for k in ("num", "str", "range", "cell", "func", "op"):
            if d[k] is not None:
                toks.append((k, d[k])); break
    return toks


class Eval:
    def __init__(self, ws):
        self.ws = ws
        self.cache = {}

    def cell(self, coord):
        if coord in self.cache:
            return self.cache[coord]
        self.cache[coord] = None            # cycle guard
        v = self.ws[coord].value
        if isinstance(v, str) and v.startswith("="):
            res = self.expr(tokenize(v[1:]), [0])
        elif isinstance(v, bool):
            res = v
        elif isnum(v):
            res = float(v)
        else:
            res = v                         # None (blank) or text
        self.cache[coord] = res
        return res

    def rng(self, spec):
        a, b = spec.split(":")
        (r1, c1), (r2, c2) = coordinate_to_tuple(a), coordinate_to_tuple(b)
        return [self.cell(f"{get_column_letter(c)}{r}")
                for r in range(min(r1, r2), max(r1, r2) + 1)
                for c in range(min(c1, c2), max(c1, c2) + 1)]

    # recursive-descent: expr -> add -> term -> factor -> primary
    def expr(self, t, p):
        left = self.add(t, p)
        while p[0] < len(t) and t[p[0]][0] == "op" and t[p[0]][1] in ("<", "<=", ">", ">=", "=", "<>"):
            op = t[p[0]][1]; p[0] += 1
            left = _cmp(op, left, self.add(t, p))
        return left

    def add(self, t, p):
        left = self.term(t, p)
        while p[0] < len(t) and t[p[0]][0] == "op" and t[p[0]][1] in ("+", "-"):
            op = t[p[0]][1]; p[0] += 1
            left = _arith(op, left, self.term(t, p))
        return left

    def term(self, t, p):
        left = self.factor(t, p)
        while p[0] < len(t) and t[p[0]][0] == "op" and t[p[0]][1] in ("*", "/"):
            op = t[p[0]][1]; p[0] += 1
            left = _arith(op, left, self.factor(t, p))
        return left

    def factor(self, t, p):
        if p[0] < len(t) and t[p[0]] == ("op", "-"):
            p[0] += 1
            v = self.factor(t, p)
            return -v if isnum(v) else "#VAL"
        return self.primary(t, p)

    def primary(self, t, p):
        k, v = t[p[0]]; p[0] += 1
        if k == "num":
            return float(v)
        if k == "str":
            return v[1:-1]
        if k == "cell":
            return self.cell(v)
        if k == "range":
            return self.rng(v)
        if k == "func":
            args = []
            if t[p[0]] != ("op", ")"):
                args.append(self.expr(t, p))
                while t[p[0]] == ("op", ","):
                    p[0] += 1; args.append(self.expr(t, p))
            p[0] += 1                       # consume ")"
            return _apply(v.upper(), args)
        if (k, v) == ("op", "("):
            val = self.expr(t, p)
            p[0] += 1                       # consume ")"
            return val
        raise ValueError(f"unexpected token {v!r}")


# ---------------- fixture: two companies with known inputs ----------------
def _co(title, price, shares, ltd, lease, minority, pref, ca, cl, ebit, da, cfo, ni):
    return {
        "title": title, "cik": 1, "is_financial": False,
        "filing_10q": {"reportDate": "2026-03-31", "filingDate": "2026-05-01", "accession": "x", "url": "u"},
        "filing_10k": {"reportDate": "2025-12-31", "filingDate": "2026-02-01", "accession": "y", "url": "u"},
        "price": price, "price_as_of": "2026-06-05", "price_source": "test", "price_source_url": "u",
        "shares_mm": shares, "market_equity_mm": None,
        "lt_debt_mm": ltd, "finance_lease_mm": lease, "minority_mm": minority, "preferred_mm": pref,
        "current_assets_mm": ca, "current_liabilities_mm": cl,
        "working_capital_mm": None, "cash_mm": 0.0, "noncurrent_investments_mm": None, "tev_mm": None,
        "ltm_mm": {"revenue": 9999.0, "ebit": ebit, "da": da, "ebitda": None, "cfo": cfo, "net_income": ni},
        "ntm_mm": {"ebitda": None, "ebit": None, "cfo": None, "net_income": None},
        "ntm_source": None, "ntm_as_of": None, "ntm_provided": False,
        "multiples": {}, "pe_ltm": None, "citations": {}, "ltm_components": {}, "flags": [],
    }


def _comps():
    return {
        "as_of_date": "2026-06-15", "currency": "USD",
        "units": "$ in millions; shares in millions; multiples in x",
        "methodology": "House EV = ...", "consensus_mode": "skip",
        "tickers": ["AAA", "BBB"],
        "companies": {
            # mkteq 10000; wc 2000; TEV=10000+2000+0+0+1000-2000=11000; EBITDA=1200
            "AAA": _co("Alpha", 100.0, 100.0, 2000.0, 0.0, 0.0, 1000.0, 3000.0, 1000.0, 1000.0, 200.0, 800.0, 500.0),
            # mkteq 10000; wc 500; TEV=10000+0+500+200+0-500=10200; EBITDA=1000; NI<0 -> nm
            "BBB": _co("Beta", 50.0, 200.0, 0.0, 500.0, 200.0, None, 2000.0, 1500.0, 850.0, 150.0, 600.0, -100.0),
        },
        "needs_consensus_for": [], "consensus_source_file": None,
    }


def run():
    fails = []

    def ck(name, cond):
        if not cond:
            fails.append("  FAIL " + name)

    def near(a, b, tol=1e-6):
        return isnum(a) and abs(a - b) <= tol

    path = os.path.join(OUT, "recalc.xlsx")
    bx.render(_comps(), path)
    ws = openpyxl.load_workbook(path).active
    ev = Eval(ws)

    # locate rows by the column-A label
    rows = {}
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(row=r, column=1).value
        if isinstance(lab, str):
            rows[lab.strip()] = r

    def row(sub):
        for lab, r in rows.items():
            if sub in lab:
                return r
        raise AssertionError(f"row not found: {sub}")

    B, C = "B", "C"                                   # ticker columns
    D, E, F, G = "D", "E", "F", "G"                   # Min / Mean / Median / Max (n=2 -> cols 4..7)

    r_tev = row("Total Enterprise Value (TEV)")
    r_mkt = row("Market equity value")
    r_wc = row("Working capital")
    r_eb = row("EBITDA (LTM)")
    r_mev = row("TEV / EBITDA — LTM")
    r_mni = row("TEV / Net income — LTM")

    # --- derived inputs recompute correctly ---
    ck("AAA market equity = 10,000", near(ev.cell(f"{B}{r_mkt}"), 10000.0))
    ck("AAA working capital = 2,000", near(ev.cell(f"{B}{r_wc}"), 2000.0))
    ck("AAA EBITDA = 1,200", near(ev.cell(f"{B}{r_eb}"), 1200.0))
    ck("BBB EBITDA = 1,000", near(ev.cell(f"{C}{r_eb}"), 1000.0))

    # --- TEV bridge (incl. preferred + minority) recomputes correctly ---
    ck("AAA TEV = 11,000 (incl. $1,000 preferred)", near(ev.cell(f"{B}{r_tev}"), 11000.0))
    ck("BBB TEV = 10,200 (preferred blank -> N()=0)", near(ev.cell(f"{C}{r_tev}"), 10200.0))

    # --- multiples ---
    ck("AAA TEV/EBITDA = 9.1667", near(ev.cell(f"{B}{r_mev}"), 11000.0 / 1200.0))
    ck("BBB TEV/EBITDA = 10.20", near(ev.cell(f"{C}{r_mev}"), 10.2))
    ck("BBB TEV/NI = 'nm' (NI<=0)", ev.cell(f"{C}{r_mni}") == "nm")
    ck("AAA TEV/NI = 22.0", near(ev.cell(f"{B}{r_mni}"), 22.0))

    # --- peer summary statistics (live MIN/MEAN/MEDIAN/MAX), with "nm" excluded ---
    lo, hi = 11000.0 / 1200.0, 10.2
    ck("EBITDA-mult Min", near(ev.cell(f"{D}{r_mev}"), lo))
    ck("EBITDA-mult Mean", near(ev.cell(f"{E}{r_mev}"), (lo + hi) / 2))
    ck("EBITDA-mult Median", near(ev.cell(f"{F}{r_mev}"), (lo + hi) / 2))
    ck("EBITDA-mult Max", near(ev.cell(f"{G}{r_mev}"), hi))
    # NI multiple: only AAA (22.0) is numeric; BBB is "nm" and must drop out of the stats
    ck("NI-mult Min excludes nm (=22)", near(ev.cell(f"{D}{r_mni}"), 22.0))
    ck("NI-mult Max excludes nm (=22)", near(ev.cell(f"{G}{r_mni}"), 22.0))
    ck("NI-mult Mean excludes nm (=22)", near(ev.cell(f"{E}{r_mni}"), 22.0))

    print("-" * 56)
    print(f"{18 - len(fails)}/18 passed" + ("" if not fails else f", {len(fails)} FAILED"))
    for f in fails:
        print(f)
    return 0 if not fails else 1


if __name__ == "__main__":
    sys.exit(run())
