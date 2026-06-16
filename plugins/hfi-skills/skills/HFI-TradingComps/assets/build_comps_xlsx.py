#!/usr/bin/env python3
"""
build_comps_xlsx.py — Render the assembled comps dict (from build_comps.assemble) into a clean,
FORMULA-DRIVEN Excel workbook so every calculation is auditable and live:

  * Filing inputs (price, shares, debt, leases, NCI, current A/L, LTM EBIT/D&A/CFO/NI) are written
    as static, sourced values (a comp is a dated snapshot).
  * Derived cells are real Excel formulas referencing the input cells:
      market equity = price x shares ;  working capital = curr assets - curr liab
      TEV = mkt equity + LT debt + finance leases + minority interest - working capital
      EBITDA = EBIT + D&A ;  every multiple = TEV / denominator (with nm/blank guards)
  * NTM consensus rows are live `=CIQ("TKR","IQ_..._EST",IQ_NTM)` Capital IQ plug-in formulas
    (consensus_mode='capiq_excel'); they populate from the user's CapIQ login on open. If a manual
    consensus value was supplied it is written statically instead; if mode='skip' the row is blank.
    NTM multiples reference the CIQ cells, so they recompute the moment consensus lands.

render(comps, path) -> path
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"; MIDBLUE = "2E5496"; GREY = "F2F2F2"; GOLD = "BF9000"
FMT_PRICE = '#,##0.00'
FMT_MM = '#,##0.0;(#,##0.0)'
FMT_X = '0.0"x";(0.0"x")'
FONT = "Calibri"
med = Side(style="medium", color="404040")

NTM_MNEMONIC = {"ebitda": "IQ_EBITDA_EST", "ebit": "IQ_EBIT_EST",
                "cfo": "IQ_CFO_EST", "net_income": "IQ_NI_EST"}


def render(comps, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trading Comps"
    tickers = comps["tickers"]
    mode = comps.get("consensus_mode", "capiq_excel")

    def col(i):
        return get_column_letter(2 + i)

    state = {"r": 1}
    R = {}  # row-key -> row number (for formula references)

    def cell(row, c, value=None, fmt=None, bold=False, italic=False, size=11,
             color=None, fill=None, align=None, border=None, wrap=False):
        cc = ws.cell(row=row, column=c)
        if value is not None:
            cc.value = value
        cc.font = Font(name=FONT, bold=bold, italic=italic, size=size, color=(color or "000000"))
        if fmt:
            cc.number_format = fmt
        if fill:
            cc.fill = PatternFill("solid", fgColor=fill)
        al = {"vertical": "center"}
        if align:
            al["horizontal"] = align
        if wrap:
            al["wrap_text"] = True
        cc.alignment = Alignment(**al)
        if border:
            cc.border = border
        return cc

    def section(label):
        r = state["r"]
        cell(r, 1, label, bold=True, color="FFFFFF", fill=MIDBLUE, size=10)
        for j in range(len(tickers)):
            cell(r, 2 + j, "", fill=MIDBLUE)
        state["r"] += 1

    def blank():
        state["r"] += 1

    def _label(r, label, bold, indent, total):
        lbl = ("    " * indent) + label
        cell(r, 1, lbl, bold=bold or total, size=10,
             border=(Border(top=med) if total else None))

    def emit_input(key, label, getter, fmt=FMT_MM, bold=False, indent=0):
        r = state["r"]; R[key] = r
        _label(r, label, bold, indent, False)
        for i, t in enumerate(tickers):
            v = getter(comps["companies"][t])
            cell(r, 2 + i, v, fmt=fmt, bold=bold, align="right")
        state["r"] += 1

    def emit_formula(key, label, fmlfn, fmt=FMT_MM, bold=False, total=False, indent=0):
        r = state["r"]; R[key] = r
        _label(r, label, bold, indent, total)
        for i in range(len(tickers)):
            cell(r, 2 + i, fmlfn(col(i)), fmt=fmt, bold=bold or total, align="right",
                 border=(Border(top=med) if total else None))
        state["r"] += 1

    def emit_ntm(key, label, metric, bold=False):
        r = state["r"]; R[key] = r
        _label(r, label, bold, 0, False)
        for i, t in enumerate(tickers):
            provided = comps["companies"][t].get("ntm_mm", {}).get(metric)
            if provided is not None:
                cell(r, 2 + i, provided, fmt=FMT_MM, align="right")
            elif mode == "capiq_excel":
                f = f'=CIQ("{t}","{NTM_MNEMONIC[metric]}",IQ_NTM)'
                cell(r, 2 + i, f, fmt=FMT_MM, align="right")
            else:
                cell(r, 2 + i, None, align="right")
        state["r"] += 1

    def mult(c, num_row, den_row):
        num, den = f"{c}{num_row}", f"{c}{den_row}"
        return (f'=IF(NOT(ISNUMBER({den})),"",IF({den}<=0,"nm",'
                f'IF(ISNUMBER({num}),{num}/{den},"")))')

    def emit_mult(key, label, num_key, den_key, bold=False):
        emit_formula(key, label, lambda c: mult(c, R[num_key], R[den_key]), fmt=FMT_X, bold=bold)

    # ---------- Title ----------
    r = state["r"]
    cell(r, 1, "Trading Statistics — Comparable Company Analysis", bold=True, size=15, color=NAVY)
    state["r"] += 1
    cell(state["r"], 1, f"Total Enterprise Value & Valuation Multiples · House methodology · "
                        f"{comps['units']} · Prices as of {comps['as_of_date']}",
         italic=True, size=9, color="595959")
    state["r"] += 1
    cell(state["r"], 1, comps["methodology"], italic=True, size=9, color="595959")
    state["r"] += 2

    # ---------- Header ----------
    hdr_row = state["r"]
    cell(hdr_row, 1, "($ in millions unless noted)", bold=True, color="FFFFFF", fill=NAVY, size=10)
    for j, t in enumerate(tickers):
        cell(hdr_row, 2 + j, t, bold=True, color="FFFFFF", fill=NAVY, align="center", size=12)
    state["r"] += 1
    cell(state["r"], 1, "", fill=GREY)
    for j, t in enumerate(tickers):
        cell(state["r"], 2 + j, comps["companies"][t]["title"], italic=True, size=8,
             color="595959", align="center", fill=GREY, wrap=True)
    ws.row_dimensions[state["r"]].height = 24
    state["r"] += 1

    # ---------- Market data ----------
    section("MARKET DATA")
    emit_input("price", "Share price ($)", lambda co: co["price"], fmt=FMT_PRICE)
    emit_input("shares", "Shares outstanding (mm)", lambda co: co["shares_mm"])
    emit_formula("mkteq", "Market equity value",
                 lambda c: f'=IF(AND(ISNUMBER({c}{R["price"]}),ISNUMBER({c}{R["shares"]})),'
                           f'{c}{R["price"]}*{c}{R["shares"]},"")', bold=True)

    # ---------- EV bridge ----------
    section("ENTERPRISE VALUE BRIDGE")
    emit_input("ltdebt", "(+) Long-term debt", lambda co: co["lt_debt_mm"])
    emit_input("lease", "(+) Finance / capital lease obligations", lambda co: co["finance_lease_mm"])
    emit_input("minority", "(+) Minority / noncontrolling interest", lambda co: co["minority_mm"])
    emit_input("preferred", "(+) Preferred equity", lambda co: co.get("preferred_mm"))
    # Working capital is shown as the main (–) line with current assets/liabs as indented sub-rows
    # beneath it, so reserve its row first, emit the sub-rows, then fill the WC formula.
    R["wc"] = state["r"]
    _label(state["r"], "(–) Working capital", False, 0, False)
    state["r"] += 1
    emit_input("ca", "Total current assets", lambda co: co["current_assets_mm"], indent=1)
    emit_input("cl", "Total current liabilities", lambda co: co["current_liabilities_mm"], indent=1)
    for i in range(len(tickers)):
        c = col(i)
        cell(R["wc"], 2 + i,
             f'=IF(AND(ISNUMBER({c}{R["ca"]}),ISNUMBER({c}{R["cl"]})),{c}{R["ca"]}-{c}{R["cl"]},"")',
             fmt=FMT_MM, align="right")
    emit_formula("tev", "Total Enterprise Value (TEV)",
                 lambda c: f'=IF(ISNUMBER({c}{R["mkteq"]}),{c}{R["mkteq"]}+N({c}{R["ltdebt"]})+'
                           f'N({c}{R["lease"]})+N({c}{R["minority"]})+N({c}{R["preferred"]})-'
                           f'N({c}{R["wc"]}),"")',
                 bold=True, total=True)
    emit_input("cash", "(memo) Cash & equivalents", lambda co: co["cash_mm"])

    # ---------- LTM ----------
    section("LTM OPERATING METRICS")
    emit_input("revenue", "Revenue (LTM)", lambda co: co["ltm_mm"]["revenue"])
    emit_input("ebit", "EBIT (LTM)", lambda co: co["ltm_mm"]["ebit"])
    emit_input("da", "(+) D&A (LTM)", lambda co: co["ltm_mm"]["da"], indent=1)
    emit_formula("ebitda", "EBITDA (LTM)",
                 lambda c: f'=IF(AND(ISNUMBER({c}{R["ebit"]}),ISNUMBER({c}{R["da"]})),'
                           f'{c}{R["ebit"]}+{c}{R["da"]},"")', bold=True)
    emit_input("cfo", "CFO (LTM)", lambda co: co["ltm_mm"]["cfo"])
    emit_input("ni", "Net income (LTM)", lambda co: co["ltm_mm"]["net_income"])

    # ---------- NTM consensus ----------
    section("NTM CONSENSUS (Wall Street via Capital IQ)")
    emit_ntm("ntm_ebitda", "EBITDA (NTM)", "ebitda")
    emit_ntm("ntm_ebit", "EBIT (NTM)", "ebit")
    emit_ntm("ntm_cfo", "CFO (NTM)", "cfo")
    emit_ntm("ntm_ni", "Net income (NTM)", "net_income")

    # ---------- Valuation ----------
    blank()
    section("VALUATION MULTIPLES (x)")
    emit_mult("m_ev_ebitda_ltm", "TEV / EBITDA — LTM", "tev", "ebitda", bold=True)
    emit_mult("m_ev_ebit_ltm", "TEV / EBIT — LTM", "tev", "ebit")
    emit_mult("m_ev_cfo_ltm", "TEV / CFO — LTM", "tev", "cfo")
    emit_mult("m_ev_ni_ltm", "TEV / Net income — LTM", "tev", "ni")
    blank()
    emit_mult("m_ev_ebitda_ntm", "TEV / EBITDA — NTM", "tev", "ntm_ebitda", bold=True)
    emit_mult("m_ev_ebit_ntm", "TEV / EBIT — NTM", "tev", "ntm_ebit")
    emit_mult("m_ev_cfo_ntm", "TEV / CFO — NTM", "tev", "ntm_cfo")
    emit_mult("m_ev_ni_ntm", "TEV / Net income — NTM", "tev", "ntm_ni")
    blank()
    emit_mult("pe_ltm", "(memo) P/E — LTM (mkt equity / NI)", "mkteq", "ni")

    # ---------- Peer summary statistics ----------
    # Min / Mean / Median / Max of each multiple ACROSS the peer set, as columns to the right of the
    # ticker columns (standard comp-sheet layout). Live formulas; Excel's COUNT/MEDIAN/MIN/MAX/AVERAGE
    # ignore text ("nm") and blanks, so non-meaningful or missing multiples drop out of the peer stats.
    n = len(tickers)
    if n >= 2:
        first_c, last_c = col(0), col(n - 1)
        stat_cols = [("Min", "MIN"), ("Mean", "AVERAGE"), ("Median", "MEDIAN"), ("Max", "MAX")]
        for j, (label, _) in enumerate(stat_cols):
            cidx = 2 + n + j
            cell(hdr_row, cidx, label, bold=True, color="FFFFFF", fill=GOLD, align="center", size=11)
            ws.column_dimensions[get_column_letter(cidx)].width = 11
        for key in ("m_ev_ebitda_ltm", "m_ev_ebit_ltm", "m_ev_cfo_ltm", "m_ev_ni_ltm",
                    "m_ev_ebitda_ntm", "m_ev_ebit_ntm", "m_ev_cfo_ntm", "m_ev_ni_ntm", "pe_ltm"):
            r = R.get(key)
            if not r:
                continue
            rng = f"{first_c}{r}:{last_c}{r}"
            for j, (label, fn) in enumerate(stat_cols):
                cell(r, 2 + n + j, f'=IF(COUNT({rng})=0,"",{fn}({rng}))',
                     fmt=FMT_X, align="right", fill=GREY, size=10,
                     bold=(key == "m_ev_ebitda_ltm"))

    # ---------- Legend + footnotes ----------
    blank()
    cell(state["r"], 1, 'Legend:  "nm" = not meaningful (denominator ≤ 0).  Blank = not available / '
                        'consensus pending.  Negatives in (parentheses).  Derived cells are live formulas.',
         italic=True, size=8, color="595959")
    state["r"] += 2

    cell(state["r"], 1, "SOURCES & FOOTNOTES — every figure traces to a primary source", bold=True,
         color="FFFFFF", fill=NAVY, size=10)
    state["r"] += 1
    notes = [
        "Methodology: TEV = market equity value + long-term debt + finance/capital lease obligations "
        "+ preferred equity + minority interest − working capital (= total current assets − total "
        "current liabilities, most recent 10-Q). Minority interest sums equity-section and redeemable "
        "(mezzanine) NCI; preferred is added at carrying value (par may understate liquidation value — "
        "see name footnotes). Only long-term debt and non-current finance leases are added; current "
        "portions sit in current liabilities and are netted via working capital (no double count). "
        "House definition — differs from textbook EV (which subtracts cash only).",
        "Summary columns (Min / Mean / Median / Max, right of the peer columns) are live formulas over "
        "each multiple row; \"nm\" and blank cells are automatically excluded from the peer statistics.",
        "Market equity value = shares outstanding (most recent 10-Q cover) × latest stock price. "
        "Derived cells (market equity, working capital, TEV, EBITDA, all multiples) are LIVE EXCEL "
        "FORMULAS referencing the input rows above — change an input and everything recomputes.",
        "LTM = latest fiscal year (10-K) + latest YTD (10-Q) − prior-year comparable YTD. "
        "EBITDA (LTM) = EBIT (operating income) + D&A. Filing data from SEC EDGAR XBRL (data.sec.gov).",
        ("NTM rows are live Capital IQ plug-in formulas, e.g. =CIQ(\"AAPL\",\"IQ_EBITDA_EST\",IQ_NTM); "
         "they populate from your CapIQ login on open. If you see #NAME?, the CapIQ Excel add-in is not "
         "loaded — enable it, paste consensus manually, or pull via CapIQ web. CFO consensus (IQ_CFO_EST) "
         "coverage is thin; verify or leave blank.") if mode == "capiq_excel" else
        ("NTM (Next-Twelve-Months) consensus is from Financial Modeling Prep analyst estimates, "
         "calendarized to a true next-12-months figure (a time-weighted blend of the current and next "
         "fiscal-year consensus). FMP does not provide a CFO estimate, so TEV/CFO-NTM stays blank. "
         "Values are static as of the run date; verify against the provider for a decision.")
        if mode == "fmp" else
        ("NTM (Next-Twelve-Months) consensus columns are optional and left blank in this run — the "
         "LTM analysis above is complete and needs no paid data subscription. To populate NTM later: "
         "re-run with --consensus-mode fmp (a free Financial Modeling Prep key), the Capital IQ Excel "
         "add-in, or paste estimates manually."),
        "For multi-class issuers the CapIQ identifier may need adjusting (e.g. \"BRK.B\" or "
        "\"NYSE:BRK.B\") and market equity should sum all classes × their prices.",
    ]
    for n in notes:
        cell(state["r"], 1, "•  " + n, size=8, wrap=True, color="404040")
        ws.row_dimensions[state["r"]].height = 30
        state["r"] += 1
    state["r"] += 1

    for t in tickers:
        co = comps["companies"][t]
        q = co.get("filing_10q") or {}
        k = co.get("filing_10k") or {}
        cell(state["r"], 1, f"{t} — {co['title']}  (CIK {co['cik']})", bold=True, size=9, color=NAVY)
        state["r"] += 1
        cell(state["r"], 1, f"   Price: {co['price']} ({co.get('price_source','')}), as of {co.get('price_as_of','')}",
             size=8, color="404040"); state["r"] += 1
        if q:
            cell(state["r"], 1, f"   10-Q: period {q.get('reportDate')}, filed {q.get('filingDate')}, "
                                f"accession {q.get('accession')} — {q.get('url')}", size=8, color="404040", wrap=True)
            ws.row_dimensions[state["r"]].height = 22; state["r"] += 1
        if k:
            cell(state["r"], 1, f"   10-K: period {k.get('reportDate')}, filed {k.get('filingDate')}, "
                                f"accession {k.get('accession')} — {k.get('url')}", size=8, color="404040", wrap=True)
            ws.row_dimensions[state["r"]].height = 22; state["r"] += 1
        for fl in co.get("flags", []):
            cell(state["r"], 1, "   ⚠ " + fl, size=8, color="C00000", wrap=True)
            ws.row_dimensions[state["r"]].height = 22; state["r"] += 1
        state["r"] += 1

    ws.column_dimensions["A"].width = 46
    for j in range(len(tickers)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 15
    ws.freeze_panes = f"B{hdr_row + 2}"
    ws.sheet_view.showGridLines = False
    wb.save(path)
    return path


if __name__ == "__main__":
    import json, sys
    with open(sys.argv[1], "r", encoding="utf-8") as f:
        comps = json.load(f)
    print(render(comps, sys.argv[2]))
