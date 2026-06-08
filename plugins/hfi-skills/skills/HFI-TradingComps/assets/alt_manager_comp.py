#!/usr/bin/env python3
"""
alt_manager_comp.py — Build a comp for ALTERNATIVE ASSET MANAGERS / financials on the metrics the
Street actually uses (AUM, Fee-Related Earnings, Distributable Earnings; P/E, P/FRE, P/Distributable,
market cap / AUM). These names (BX, KKR, ARES, OWL, APO, CG, TPG, ...) do NOT fit the EV/EBITDA house
methodology — see reference/financials_and_alts.md.

Data is NOT in standard XBRL; hand-extract each figure from the firm's quarterly earnings release
(use fetch_earnings.py) and fill the per-company dict, then call render(data, out_path).

render(data, out_path, as_of="YYYY-MM-DD") -> out_path
  data = {TICKER: {name, price, class_a, econ, aum, fpaum, fre, dist, dist_name, ni,
                   fre_ps, dist_ps, eps, source}}   # $M; AUM in $B; per-share in $
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, MID, GREY = "1F3864", "2E5496", "F2F2F2"
FONT = "Calibri"
MED = Side(style="medium", color="404040")


def render(data, out_path, as_of="(date)"):
    tick = list(data.keys())
    for t in tick:
        c = data[t]
        c["ca_mktcap"] = c["class_a"] * c["price"]
        c["econ_mktcap"] = c["econ"] * c["price"]
        c["pe"] = (c["ca_mktcap"] / c["ni"]) if c.get("ni") else None
        c["p_fre"] = (c["price"] / c["fre_ps"]) if c.get("fre_ps") else None
        c["p_dist"] = (c["price"] / c["dist_ps"]) if c.get("dist_ps") else None
        c["cap_aum"] = (c["econ_mktcap"] / (c["aum"] * 1000.0) * 100.0) if c.get("aum") else None

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Alt Manager Comp"
    st = {"r": 1}

    def cell(row, col, v=None, fmt=None, bold=False, it=False, sz=11, color=None, fill=None, al=None, bd=None, wrap=False):
        cc = ws.cell(row=row, column=col)
        if v is not None: cc.value = v
        cc.font = Font(name=FONT, bold=bold, italic=it, size=sz, color=(color or "000000"))
        if fmt: cc.number_format = fmt
        if fill: cc.fill = PatternFill("solid", fgColor=fill)
        a = {"vertical": "center"}
        if al: a["horizontal"] = al
        if wrap: a["wrap_text"] = True
        cc.alignment = Alignment(**a)
        if bd: cc.border = bd
        return cc

    cell(st["r"], 1, "Alternative Asset Managers — Comparable Company Analysis", bold=True, sz=15, color=NAVY); st["r"] += 1
    cell(st["r"], 1, f"Valued on AUM / Fee-Related Earnings / Distributable Earnings (NOT EV/EBITDA). "
                     f"$ in millions unless noted · AUM in $B · Prices as of {as_of} · LTM where shown",
         it=True, sz=9, color="595959"); st["r"] += 2
    hdr = st["r"]
    cell(hdr, 1, "($M unless noted)", bold=True, color="FFFFFF", fill=NAVY, sz=10)
    for j, t in enumerate(tick): cell(hdr, 2 + j, t, bold=True, color="FFFFFF", fill=NAVY, al="center", sz=12)
    st["r"] += 1
    cell(st["r"], 1, "", fill=GREY)
    for j, t in enumerate(tick): cell(st["r"], 2 + j, data[t]["name"], it=True, sz=8, color="595959", al="center", fill=GREY)
    st["r"] += 1

    def sec(lbl):
        cell(st["r"], 1, lbl, bold=True, color="FFFFFF", fill=MID, sz=10)
        for j in range(len(tick)): cell(st["r"], 2 + j, "", fill=MID)
        st["r"] += 1

    def row(lbl, key, fmt='#,##0.0', bold=False, total=False):
        cell(st["r"], 1, lbl, bold=bold or total, sz=10, bd=(Border(top=MED) if total else None))
        for j, t in enumerate(tick):
            cell(st["r"], 2 + j, data[t].get(key), fmt=fmt, bold=bold or total, al="right",
                 bd=(Border(top=MED) if total else None))
        st["r"] += 1

    sec("MARKET DATA")
    row("Share price ($)", "price", fmt='#,##0.00')
    row("Class A shares (mm)", "class_a")
    row("Economic / adjusted shares (mm)", "econ")
    row("Class A market cap", "ca_mktcap", bold=True)
    row("Economic market cap (incl. Op-Group units)", "econ_mktcap", bold=True)
    sec("SCALE")
    row("Assets Under Management ($B)", "aum")
    row("Fee-Earning / Fee-Paying AUM ($B)", "fpaum")
    sec("EARNINGS (LTM)")
    row("Fee-Related Earnings (FRE)", "fre")
    row("Distributable / realized earnings*", "dist")
    row("GAAP net income to public co.", "ni")
    sec("PER SHARE (LTM, $)")
    row("FRE per share", "fre_ps", fmt='#,##0.00')
    row("Distributable per share*", "dist_ps", fmt='#,##0.00')
    row("GAAP diluted EPS", "eps", fmt='#,##0.00')
    sec("VALUATION")
    row("P / E (GAAP)", "pe", fmt='0.0"x"', bold=True)
    row("P / FRE", "p_fre", fmt='0.0"x"', bold=True)
    row("P / Distributable*", "p_dist", fmt='0.0"x"', bold=True)
    row("Economic mkt cap / AUM (%)", "cap_aum", fmt='0.0"%"')
    st["r"] += 1
    cell(st["r"], 1, "* Distributable metric differs by firm — see footnotes (BX/OWL Distributable Earnings; "
                     "ARES After-tax Realized Income; KKR Adjusted Net Income). GAAP P/E is distorted by Up-C "
                     "structures (most economics accrue to operating-group units / NCI) — use FRE & Distributable.",
         it=True, sz=8, color="595959", wrap=True); ws.row_dimensions[st["r"]].height = 28; st["r"] += 2
    cell(st["r"], 1, "SOURCES (primary — each firm's quarterly earnings release + 10-Q)", bold=True,
         color="FFFFFF", fill=NAVY, sz=10); st["r"] += 1
    for t in tick:
        cell(st["r"], 1, f"{t} — {data[t]['name']}", bold=True, sz=9, color=NAVY); st["r"] += 1
        cell(st["r"], 1, "   " + data[t].get("source", ""), sz=8, color="404040", wrap=True)
        ws.row_dimensions[st["r"]].height = 34; st["r"] += 1

    ws.column_dimensions["A"].width = 44
    for j in range(len(tick)): ws.column_dimensions[get_column_letter(2 + j)].width = 15
    ws.freeze_panes = f"B{hdr + 2}"; ws.sheet_view.showGridLines = False
    wb.save(out_path)
    return out_path


# ---- Worked example: OWL / ARES / BX / KKR, Q1 2026 (primary-sourced) ----
EXAMPLE = {
 "OWL": dict(name="Blue Owl Capital", price=9.80, class_a=675.802, econ=1559.0, aum=314.9, fpaum=188.4,
   fre=1544.8, dist=1339.1, dist_name="Distributable Earnings", ni=313.6, fre_ps=0.99, dist_ps=0.86, eps=0.46,
   source="8-K 2026-04-30 (0001823945-26-000022) EX-99.1+deck; 10-Q 2026-03-31. AUM $314.9B / FPAUM $188.4B; FRE LTM $1,544.8M; DE LTM $1,339.1M; Class A 675,802,413 (+~883M Op Group units)."),
 "ARES": dict(name="Ares Management", price=125.65, class_a=222.028, econ=329.85, aum=644.3, fpaum=399.6,
   fre=1857.6, dist=1775.2, dist_name="After-tax Realized Income", ni=521.5, fre_ps=5.63, dist_ps=5.38, eps=2.33,
   source="8-K 2026-05-01 (0001628280-26-029083) EX-99.1/99.2; 10-Q. AUM $644.3B / FPAUM $399.6B; FRE Q1 $464.4M (LTM annualized); ATRI LTM $1,775.2M; Class A 222,028,421 (+3.5M non-voting, +104.3M AOG units)."),
 "BX": dict(name="Blackstone", price=115.35, class_a=742.880, econ=1230.17, aum=1304.0, fpaum=937.6,
   fre=6023.5, dist=7500.0, dist_name="Distributable Earnings", ni=3054.1, fre_ps=4.90, dist_ps=5.84, eps=3.90,
   source="8-K 2026-04-23 (0001193125-26-171788) EX-99.1. AUM $1,304.0B / Fee-Earning $937.6B; FRE LTM $6,023.5M ($4.90/sh); DE LTM ~$7.5B ($5.84/sh); GAAP NI LTM $3,054.1M; Class A 742.9M / DE shares 1,230.2M."),
 "KKR": dict(name="KKR & Co.", price=93.40, class_a=889.414, econ=900.08, aum=757.9, fpaum=614.8,
   fre=3906.0, dist=4590.0, dist_name="Adjusted Net Income", ni=2802.6, fre_ps=4.34, dist_ps=5.10, eps=2.93,
   source="8-K 2026-05-05 (0001404912-26-000011) release. AUM $757.9B / FPAUM $614.8B; FRE LTM $3.9B ($4.34/adj sh); ANI LTM ($5.10/adj sh); GAAP NI LTM $2,802.6M; common 889.4M / adjusted 900.1M."),
}

if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "Alt_Manager_Comp_example.xlsx"
    print("Wrote:", render(EXAMPLE, out, as_of="2026-06-05"))
